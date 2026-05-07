#!/usr/bin/env python3
"""
Convert Insight support matrix CSV to formatted Excel file.

Usage:
    python tools/render2xlsx.py <input.csv> [--output <output.xlsx>]

Example:
    python tools/render2xlsx.py test_cpu_support.csv
    python tools/render2xlsx.py test_cpu_support.csv -o support_matrix.xlsx
"""

import csv
import argparse
from pathlib import Path

# Try to import openpyxl, guide user if not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("Error: openpyxl not installed.")
    print("Install with: pip install openpyxl")
    import sys
    sys.exit(1)


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """Convert CSV to formatted Excel file."""
    
    # Read CSV
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        data = list(reader)
    
    if not data:
        print("Error: CSV file is empty")
        sys.exit(1)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Support Matrix"
    
    # Write data
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            # Convert "1"/"0" to integers for numeric cells, keep header and operator names as text
            if row_idx > 1 and col_idx > 1:
                # Value should be '1' or '0' -> store as int to avoid green triangle
                cell_value = int(value) if value in ('0', '1') else value
            else:
                cell_value = value
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # ========== Apply formatting ==========
    
    # Header styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Support styles (value = 1)
    support_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    support_font = Font(color="006100")
    
    # Unsupported styles (value = 0)
    unsupport_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    unsupport_font = Font(color="9C0006")
    
    # Format cells
    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if row_idx == 1:
                # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            elif col_idx == 1:
                # Operator name column
                cell.font = Font(bold=True)
            else:
                # Data cells (1 or 0 as integers now)
                if value == "1":
                    cell.fill = support_fill
                    cell.font = support_font
                elif value == "0":
                    cell.fill = unsupport_fill
                    cell.font = unsupport_font
    
    # Adjust column widths
    for col_idx in range(1, len(data[0]) + 1):
        max_len = 0
        for row_idx in range(1, len(data) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        # Convert column number to letter (supports up to 26 columns, extend if needed)
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(64 + (col_idx - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 25)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(xlsx_path)
    print(f"✓ Generated: {xlsx_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert Insight support matrix CSV to formatted Excel file."
    )
    parser.add_argument(
        "csv_path",
        help="Path to input CSV file (e.g., test_cpu_support.csv)"
    )
    parser.add_argument(
        "-o", "--output",
        dest="xlsx_path",
        help="Path to output Excel file (default: same name as CSV with .xlsx extension)"
    )
    
    args = parser.parse_args()
    
    csv_path = args.csv_path
    if not Path(csv_path).exists():
        print(f"Error: CSV file not found: {csv_path}")
        sys.exit(1)
    
    if args.xlsx_path:
        xlsx_path = args.xlsx_path
    else:
        xlsx_path = Path(csv_path).with_suffix(".xlsx")
    
    csv_to_xlsx(csv_path, xlsx_path)


if __name__ == "__main__":
    main()